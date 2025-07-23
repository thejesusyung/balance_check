from __future__ import annotations

import hashlib
from io import BytesIO
from openpyxl.utils import get_column_letter
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Tuple
import sys
import logging

# Ensure imports work when running this file directly with Streamlit or Python.
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

import pandas as pd
import streamlit as st

logger = logging.getLogger(__name__)

from src.core.highlight import cells_to_highlight
from src.core.reconcile import reconcile
from src.io.loader import infer_engine
from src.io.writer import write_coloured
from src.llm import detector
from src.llm.schema import Detection


@st.cache_data
def _detect_cached(file_hash: str, content: bytes, name: str, key: str) -> Detection:
    """Run column detection with caching by file hash."""
    engine = infer_engine(name)
    df = pd.read_excel(BytesIO(content), engine=engine)
    return detector.detect_columns(df, api_key=key)


def _load_file(upload: st.runtime.uploaded_file_manager.UploadedFile) -> Tuple[pd.DataFrame, str, bytes]:
    """Save uploaded file to disk and read it into a DataFrame."""
    logger.info("Loading file %s", upload.name)
    data = upload.getvalue()
    suffix = Path(upload.name).suffix
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(data)
        path = tmp.name
    engine = infer_engine(upload.name)
    df = pd.read_excel(BytesIO(data), engine=engine)
    return df, path, data


def _run_reconcile(
    left_file: st.runtime.uploaded_file_manager.UploadedFile,
    right_file: st.runtime.uploaded_file_manager.UploadedFile,
    api_key: str,
) -> Tuple[bool, str, str, str]:
    """Process two uploads and perform reconciliation."""
    logger.info("Running reconciliation")
    detail_logger = logging.getLogger("balance_check.sum")
    if not detail_logger.handlers:
        handler = logging.FileHandler("sum.log")
        handler.setFormatter(
            logging.Formatter("%(asctime)s %(levelname)s %(message)s")
        )
        detail_logger.addHandler(handler)
        detail_logger.setLevel(logging.DEBUG)

    df_left, path_left, bytes_left = _load_file(left_file)
    df_right, path_right, bytes_right = _load_file(right_file)

    det_left = _detect_cached(hashlib.md5(bytes_left).hexdigest(), bytes_left, left_file.name, api_key)
    det_right = _detect_cached(hashlib.md5(bytes_right).hexdigest(), bytes_right, right_file.name, api_key)


    logger.info("Detected columns - left: %s", det_left.model_dump())
    logger.info("Detected columns - right: %s", det_right.model_dump())

    left_debit = pd.to_numeric(
        df_left.iloc[:, det_left.debit_column], errors="coerce"
    ).fillna(0)
    left_credit = pd.to_numeric(
        df_left.iloc[:, det_left.credit_column], errors="coerce"
    ).fillna(0)
    right_debit = pd.to_numeric(
        df_right.iloc[:, det_right.debit_column], errors="coerce"
    ).fillna(0)
    right_credit = pd.to_numeric(
        df_right.iloc[:, det_right.credit_column], errors="coerce"
    ).fillna(0)

    left_debit_letter = get_column_letter(det_left.debit_column + 1)
    for idx, val in enumerate(left_debit, start=2):
        cell = f"{left_debit_letter}{idx}"
        detail_logger.debug("Left %s: %s=%.2f", left_file.name, cell, val)
    left_debit_total = left_debit.sum()
    detail_logger.debug("Left debit total=%.2f", left_debit_total)

    left_credit_letter = get_column_letter(det_left.credit_column + 1)
    for idx, val in enumerate(left_credit, start=2):
        cell = f"{left_credit_letter}{idx}"
        detail_logger.debug("Left %s: %s=%.2f", left_file.name, cell, val)
    left_credit_total = left_credit.sum()
    detail_logger.debug("Left credit total=%.2f", left_credit_total)

    right_debit_letter = get_column_letter(det_right.debit_column + 1)
    for idx, val in enumerate(right_debit, start=2):
        cell = f"{right_debit_letter}{idx}"
        detail_logger.debug("Right %s: %s=%.2f", right_file.name, cell, val)
    right_debit_total = right_debit.sum()
    detail_logger.debug("Right debit total=%.2f", right_debit_total)

    right_credit_letter = get_column_letter(det_right.credit_column + 1)
    for idx, val in enumerate(right_credit, start=2):
        cell = f"{right_credit_letter}{idx}"
        detail_logger.debug("Right %s: %s=%.2f", right_file.name, cell, val)
    right_credit_total = right_credit.sum()
    detail_logger.debug("Right credit total=%.2f", right_credit_total)

    logger.info(
        "Early check totals - debit left vs credit right: %.2f vs %.2f, credit left vs debit right: %.2f vs %.2f",
        left_debit_total,
        right_credit_total,
        left_credit_total,
        right_debit_total,
    )

    if round(left_debit_total, 2) == round(right_credit_total, 2) and round(left_credit_total, 2) == round(right_debit_total, 2):
        out_left = write_coloured(df_left, set(), path_left)
        out_right = write_coloured(df_right, set(), path_right)
        report = (
            f"Debit total left {left_debit_total:.2f} matches credit total right {right_credit_total:.2f}\n"
            f"Credit total left {left_credit_total:.2f} matches debit total right {right_debit_total:.2f}"
        )
        logger.info("Cross totals match - skipping detailed reconciliation")
        return True, report, out_left, out_right

    matches, partials, unmatched = reconcile(df_left, df_right, det_left, det_right)

    left_cells = cells_to_highlight(matches, partials, unmatched, det_left, "left")
    right_cells = cells_to_highlight(matches, partials, unmatched, det_right, "right")

    out_left = write_coloured(df_left, left_cells, path_left)
    out_right = write_coloured(df_right, right_cells, path_right)

    success = not partials and not unmatched and all(m.diff == 0 for m in matches)
    report = f"Matches: {len(matches)}\nPartials: {len(partials)}\nUnmatched: {len(unmatched)}"

    logger.info("Reconciliation result: %s", report.replace("\n", "; "))
    return success, report, out_left, out_right


def main() -> None:
    """Streamlit entrypoint."""
    logging.basicConfig(level=logging.INFO)
    st.title("Balance Check")

    key = st.sidebar.text_input("OpenAI API Key", type="password")
    if key:
        st.session_state["openai_key"] = key

    left = st.file_uploader("Left workbook", type=["xls", "xlsx"], key="left")
    right = st.file_uploader("Right workbook", type=["xls", "xlsx"], key="right")

    if st.button("Reconcile", disabled=not (left and right)) and left and right:
        with st.spinner("Reconciling..."):
            success, report, out_left, out_right = _run_reconcile(
                left, right, st.session_state.get("openai_key", "")
            )
        if success:
            st.success("All rows matched across workbooks.")
        else:
            st.error(f"Differences found:\n{report}")
        with open(out_left, "rb") as f:
            st.download_button("Download left result", f, file_name=Path(out_left).name)
        with open(out_right, "rb") as f:
            st.download_button("Download right result", f, file_name=Path(out_right).name)
        st.text_area("Report", report, height=120)


if __name__ == "__main__":
    main()
