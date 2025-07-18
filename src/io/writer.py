"""Excel writing utilities with highlighting support."""

from __future__ import annotations

from pathlib import Path
from typing import Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.reader import drawings
from warnings import warn
from shutil import copy2


def _load_workbook_safe(path: Path):
    """Load workbook ignoring invalid drawing relationships."""
    orig_find_images = drawings.find_images

    def safe_find_images(archive, rel_path):
        try:
            return orig_find_images(archive, rel_path)
        except KeyError as exc:  # pragma: no cover - depends on corrupt input
            warn(f"Skipping broken drawing reference: {exc}")
            return [], []

    drawings.find_images = safe_find_images
    try:
        return load_workbook(path)
    finally:
        drawings.find_images = orig_find_images


def write_coloured(df: pd.DataFrame, highlights: Set[Tuple[int, int]], target_path: str) -> str:
    """Clone the workbook at ``target_path`` and apply highlights.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame that was originally read from ``target_path``. It is only
        used to determine the offset of data rows when applying highlights.
    highlights : set[tuple[int, int]]
        Set of ``(row, column)`` pairs (0-based, relative to the DataFrame)
        that should be filled with a red colour.
    target_path : str
        Path to the original workbook.

    Returns
    -------
    str
        Path to the newly written workbook with ``_checked`` suffix.
    """

    src = Path(target_path)
    if not src.exists():
        raise FileNotFoundError(target_path)

    dst = src.with_name(f"{src.stem}_checked.xlsx")
    copy2(src, dst)

    wb = _load_workbook_safe(dst)
    ws = wb.active

    fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    row_offset = 1  # account for header row written by pandas when reading
    for r, c in highlights:
        cell = ws.cell(row=r + 1 + row_offset, column=c + 1)
        cell.fill = fill

    wb.save(dst)
    return str(dst)
